#!/usr/bin/env python3
"""Spreadsheet metadata and per-sheet analysis. Always run this first to triage an XLSX/XLS/CSV file."""

import sys
import os
import json
import csv
import re
from collections import Counter


def load_workbook(path: str):
    """Load a workbook, handling xlsx, xlsm (openpyxl) and xls (xlrd)."""
    ext = os.path.splitext(path)[1].lower()

    if ext in ('.csv', '.tsv'):
        return None  # Handled separately

    elif ext in ('.xlsx', '.xlsm'):
        import openpyxl
        try:
            return openpyxl.load_workbook(path, read_only=True, data_only=False), 'openpyxl'
        except Exception:
            import openpyxl
            return openpyxl.load_workbook(path, read_only=True, data_only=True), 'openpyxl_values'

    elif ext == '.xls':
        import xlrd
        return xlrd.open_workbook(path), 'xlrd'

    else:
        print(f"ERROR: Unsupported file type: {ext}", file=sys.stderr)
        sys.exit(1)


def analyze_xlsx(wb, path: str) -> dict:
    """Analyze an openpyxl workbook."""
    sheets = []
    total_rows = 0
    total_formulas = 0

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        row_count = ws.max_row or 0
        col_count = ws.max_column or 0
        total_rows += row_count

        # Extract headers from first row
        headers = []
        if row_count >= 1:
            for col in range(1, col_count + 1):
                cell = ws.cell(row=1, column=col)
                headers.append(str(cell.value) if cell.value is not None else "")

        # Sample first 50 rows for data type analysis
        typed_rows = min(row_count, 50)
        col_types = {}
        formula_count = 0

        for row in range(1, typed_rows + 1):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    col_types.setdefault(col, Counter())['null'] += 1
                elif isinstance(cell.value, bool):
                    col_types.setdefault(col, Counter())['bool'] += 1
                elif isinstance(cell.value, int):
                    col_types.setdefault(col, Counter())['int'] += 1
                elif isinstance(cell.value, float):
                    col_types.setdefault(col, Counter())['float'] += 1
                elif isinstance(cell.value, str):
                    s = cell.value
                    if re.match(r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$', s):
                        col_types.setdefault(col, Counter())['date_str'] += 1
                    else:
                        col_types.setdefault(col, Counter())['str'] += 1
                else:
                    col_types.setdefault(col, Counter())['other'] += 1

                # Check if it's a formula (only with data_only=False)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1

        # Convert column type counters to dominant type
        col_summary = {}
        for col, counter in sorted(col_types.items()):
            dominant = counter.most_common(1)[0][0] if counter else "unknown"
            col_summary[str(col)] = {
                "header": headers[col - 1] if col <= len(headers) else "",
                "dominant_type": dominant,
                "distribution": dict(counter.most_common()),
            }

        # Merged cells (not available in read_only mode)
        try:
            merged_count = len(ws.merged_cells.ranges)
        except AttributeError:
            merged_count = 0

        sheets.append({
            "sheet": ws_name,
            "rows": row_count,
            "columns": col_count,
            "headers": headers[:20],  # Truncate header list
            "column_types": col_summary,
            "merged_cell_ranges": merged_count,
            "formula_count_sampled": formula_count,
        })

        # Count formulas more thoroughly (check first column for = prefix)
        if formula_count == 0:
            for row in range(1, min(row_count + 1, 200)):
                cell = ws.cell(row=row, column=1)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    total_formulas += 1

    wb.close()

    result = {
        "file": path,
        "format": "xlsx",
        "sheet_count": len(sheets),
        "sheet_names": [s["sheet"] for s in sheets],
        "total_rows": total_rows,
        "total_formulas_sampled": total_formulas,
        "sheets": sheets,
    }
    return result


def analyze_xlrd(wb, path: str) -> dict:
    """Analyze an xlrd workbook (legacy .xls)."""
    sheets = []
    total_rows = 0

    for ws_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(ws_idx)
        row_count = ws.nrows
        col_count = ws.ncols
        total_rows += row_count

        # Headers from first row
        headers = []
        if row_count >= 1:
            for col in range(col_count):
                headers.append(str(ws.cell_value(0, col)))

        # Sample first 50 rows for types
        typed_rows = min(row_count, 50)
        col_types = {}

        for row in range(1, typed_rows):
            for col in range(col_count):
                val = ws.cell_value(row, col)
                if val == '' or val is None:
                    col_types.setdefault(col, Counter())['null'] += 1
                elif isinstance(val, bool):
                    col_types.setdefault(col, Counter())['bool'] += 1
                elif isinstance(val, (int, float)):
                    col_types.setdefault(col, Counter())['number'] += 1
                elif isinstance(val, str):
                    col_types.setdefault(col, Counter())['str'] += 1
                else:
                    col_types.setdefault(col, Counter())['other'] += 1

        col_summary = {}
        for col, counter in sorted(col_types.items()):
            dominant = counter.most_common(1)[0][0] if counter else "unknown"
            col_summary[str(col)] = {
                "header": headers[col] if col < len(headers) else "",
                "dominant_type": dominant,
                "distribution": dict(counter.most_common()),
            }

        sheets.append({
            "sheet": ws.name,
            "rows": row_count,
            "columns": col_count,
            "headers": headers[:20],
            "column_types": col_summary,
        })

    result = {
        "file": path,
        "format": "xls",
        "sheet_count": len(sheets),
        "sheet_names": [s["sheet"] for s in sheets],
        "total_rows": total_rows,
        "total_formulas_sampled": 0,  # xlrd doesn't expose formulas easily
        "sheets": sheets,
    }
    return result


def analyze_csv(path: str) -> dict:
    """Analyze a CSV/TSV file."""
    ext = os.path.splitext(path)[1].lower()
    delimiter = '\t' if ext == '.tsv' else ','

    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)

    row_count = len(rows)
    col_count = max(len(r) for r in rows) if rows else 0
    headers = rows[0] if rows else []

    # Type analysis on first 50 data rows
    sample = rows[1:51] if len(rows) > 1 else []
    col_types = {}
    for row in sample:
        for col, val in enumerate(row):
            if val == '' or val is None:
                col_types.setdefault(col, Counter())['null'] += 1
            elif re.match(r'^-?\d+$', val.strip()):
                col_types.setdefault(col, Counter())['int'] += 1
            elif re.match(r'^-?\d+\.\d+$', val.strip()):
                col_types.setdefault(col, Counter())['float'] += 1
            elif re.match(r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$', val.strip()):
                col_types.setdefault(col, Counter())['date'] += 1
            else:
                col_types.setdefault(col, Counter())['str'] += 1

    col_summary = {}
    for col, counter in sorted(col_types.items()):
        dominant = counter.most_common(1)[0][0] if counter else "unknown"
        col_summary[str(col)] = {
            "header": headers[col] if col < len(headers) else "",
            "dominant_type": dominant,
            "distribution": dict(counter.most_common()),
        }

    return {
        "file": path,
        "format": ext.lstrip('.'),
        "sheet_count": 1,
        "sheet_names": [os.path.basename(path)],
        "total_rows": row_count,
        "total_formulas_sampled": 0,
        "sheets": [{
            "sheet": os.path.basename(path),
            "rows": row_count,
            "columns": col_count,
            "headers": headers[:20],
            "column_types": col_summary,
        }],
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: xlsx_info.py <path>", file=sys.stderr)
        sys.exit(1)

    path = sys.argv[1]
    ext = os.path.splitext(path)[1].lower()

    if ext in ('.csv', '.tsv'):
        result = analyze_csv(path)
    else:
        wb, engine = load_workbook(path)
        if engine == 'openpyxl' or engine == 'openpyxl_values':
            result = analyze_xlsx(wb, path)
        else:
            result = analyze_xlrd(wb, path)

    print(json.dumps(result, indent=2))
