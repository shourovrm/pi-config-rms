#!/usr/bin/env python3
"""Extract spreadsheet data as structured output (CSV/TSV). Supports xlsx, xlsm, xls, csv, tsv."""

import sys
import os
import csv
import io
import argparse


def extract_xlsx(path: str, sheets: list[str] | None, max_rows: int | None,
                 delimiter: str, formulas: bool) -> None:
    """Extract from xlsx/xlsm via openpyxl."""
    import openpyxl

    data_only = not formulas
    wb = openpyxl.load_workbook(path, read_only=True, data_only=data_only)
    available = wb.sheetnames

    targets = sheets if sheets else available

    for ws_name in targets:
        if ws_name not in available:
            print(f"WARNING: Sheet '{ws_name}' not found, skipping", file=sys.stderr)
            continue

        ws = wb[ws_name]
        print(f"\n=== Sheet: {ws_name} ({ws.max_row} rows × {ws.max_column} cols) ===")

        writer = csv.writer(sys.stdout, delimiter=delimiter)
        written = 0
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(v) if v is not None else '' for v in row])
            written += 1
            if max_rows and written >= max_rows:
                print(f"... (truncated at {max_rows} rows)")
                break

    wb.close()


def extract_xls(path: str, sheets: list[str] | None, max_rows: int | None,
                delimiter: str) -> None:
    """Extract from legacy .xls via xlrd."""
    import xlrd

    wb = xlrd.open_workbook(path)
    available = [wb.sheet_by_index(i).name for i in range(wb.nsheets)]
    targets = sheets if sheets else available

    for ws_name in targets:
        if ws_name not in available:
            print(f"WARNING: Sheet '{ws_name}' not found, skipping", file=sys.stderr)
            continue

        ws = wb.sheet_by_name(ws_name)
        print(f"\n=== Sheet: {ws_name} ({ws.nrows} rows × {ws.ncols} cols) ===")

        writer = csv.writer(sys.stdout, delimiter=delimiter)
        for row_idx in range(ws.nrows):
            row = [str(ws.cell_value(row_idx, col)) if ws.cell_value(row_idx, col) != '' else ''
                   for col in range(ws.ncols)]
            writer.writerow(row)
            if max_rows and row_idx + 1 >= max_rows:
                print(f"... (truncated at {max_rows} rows)")
                break


def extract_csv(path: str, max_rows: int | None, delimiter: str) -> None:
    """Extract from CSV/TSV files (detect delimiter from extension)."""
    ext = os.path.splitext(path)[1].lower()
    source_delimiter = '\t' if ext == '.tsv' else ','

    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f, delimiter=source_delimiter)
        writer = csv.writer(sys.stdout, delimiter=delimiter)

        print(f"\n=== File: {os.path.basename(path)} ===")
        written = 0
        for row in reader:
            writer.writerow(row)
            written += 1
            if max_rows and written >= max_rows:
                print(f"... (truncated at {max_rows} rows)")
                break


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extract spreadsheet data as CSV")
    parser.add_argument("path", help="Path to spreadsheet file")
    parser.add_argument("--sheets", default=None,
                        help="Comma-separated sheet names (default: all)")
    parser.add_argument("--max-rows", type=int, default=None,
                        help="Maximum rows per sheet")
    parser.add_argument("--delimiter", default=",",
                        help="Output delimiter (default: ,)")
    parser.add_argument("--formulas", action="store_true",
                        help="Show formulas instead of computed values (xlsx only)")
    parser.add_argument("--list-sheets", action="store_true",
                        help="Only list sheet names, don't extract data")
    args = parser.parse_args()

    if args.list_sheets:
        ext = os.path.splitext(args.path)[1].lower()
        if ext in ('.csv', '.tsv'):
            print(os.path.basename(args.path))
        elif ext in ('.xlsx', '.xlsm'):
            import openpyxl
            wb = openpyxl.load_workbook(args.path, read_only=True)
            for name in wb.sheetnames:
                print(name)
            wb.close()
        elif ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(args.path)
            for i in range(wb.nsheets):
                print(wb.sheet_by_index(i).name)
        sys.exit(0)

    ext = os.path.splitext(args.path)[1].lower()
    sheets_list = args.sheets.split(",") if args.sheets else None

    if ext in ('.csv', '.tsv'):
        extract_csv(args.path, args.max_rows, args.delimiter)
    elif ext in ('.xlsx', '.xlsm'):
        extract_xlsx(args.path, sheets_list, args.max_rows, args.delimiter, args.formulas)
    elif ext == '.xls':
        extract_xls(args.path, sheets_list, args.max_rows, args.delimiter)
    else:
        print(f"ERROR: Unsupported format: {ext}", file=sys.stderr)
        sys.exit(1)
