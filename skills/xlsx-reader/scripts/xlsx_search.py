#!/usr/bin/env python3
"""Search within spreadsheet cells by keyword or regex. Searches all sheets."""

import sys
import os
import re
import argparse


def search_xlsx(path: str, pattern, context_cols: int, max_matches: int) -> int:
    """Search an xlsx/xlsm file."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    total_matches = 0
    MAX_PREVIEW = 150

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        # We need to iterate all rows to search — read_only mode helps with memory
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                if cell_value is None:
                    continue
                text = str(cell_value)
                if pattern.search(text):
                    total_matches += 1
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    print(f"\n=== Sheet '{ws_name}', {col_letter}{row_idx} ===")

                    # Show context: surrounding cells in the same row
                    start_col = max(1, col_idx - context_cols)
                    end_col = min(len(row), col_idx + context_cols)
                    for c in range(start_col - 1, end_col):
                        c_one = c + 1
                        cl = openpyxl.utils.get_column_letter(c_one)
                        val = str(row[c]) if row[c] is not None else ""
                        marker = ">>>" if c_one == col_idx else "   "
                        print(f"{marker} {cl}: {val[:MAX_PREVIEW]}")

                    if total_matches >= max_matches:
                        print(f"\n--- Reached match limit ({max_matches}), stopping ---")
                        wb.close()
                        return total_matches

    wb.close()
    return total_matches


def search_xls(path: str, pattern, context_cols: int, max_matches: int) -> int:
    """Search a legacy .xls file."""
    import xlrd

    wb = xlrd.open_workbook(path)
    total_matches = 0
    MAX_PREVIEW = 150

    for ws_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(ws_idx)
        for row_idx in range(ws.nrows):
            for col_idx in range(ws.ncols):
                val = ws.cell_value(row_idx, col_idx)
                if val == '' or val is None:
                    continue
                text = str(val)
                if pattern.search(text):
                    total_matches += 1
                    # Column letter for xls
                    col_letter = ""
                    n = col_idx
                    while n >= 0:
                        col_letter = chr(ord('A') + (n % 26)) + col_letter
                        n = n // 26 - 1

                    print(f"\n=== Sheet '{ws.name}', {col_letter}{row_idx + 1} ===")

                    start_col = max(0, col_idx - context_cols)
                    end_col = min(ws.ncols, col_idx + context_cols + 1)
                    for c in range(start_col, end_col):
                        cl = ""
                        n = c
                        while n >= 0:
                            cl = chr(ord('A') + (n % 26)) + cl
                            n = n // 26 - 1
                        val = str(ws.cell_value(row_idx, c))
                        marker = ">>>" if c == col_idx else "   "
                        print(f"{marker} {cl}: {val[:MAX_PREVIEW]}")

                    if total_matches >= max_matches:
                        print(f"\n--- Reached match limit ({max_matches}), stopping ---")
                        return total_matches

    return total_matches


def search_csv(path: str, pattern, context_cols: int, max_matches: int) -> int:
    """Search within a CSV/TSV file."""
    import csv

    ext = os.path.splitext(path)[1].lower()
    delimiter = '\t' if ext == '.tsv' else ','

    total_matches = 0
    MAX_PREVIEW = 150

    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, val in enumerate(row):
                if not val or not val.strip():
                    continue
                if pattern.search(val):
                    total_matches += 1
                    col_letter = ""
                    n = col_idx
                    while n >= 0:
                        col_letter = chr(ord('A') + (n % 26)) + col_letter
                        n = n // 26 - 1

                    print(f"\n=== Row {row_idx}, {col_letter}{row_idx} ===")

                    start_col = max(0, col_idx - context_cols)
                    end_col = min(len(row), col_idx + context_cols + 1)
                    for c in range(start_col, end_col):
                        cl = ""
                        n = c
                        while n >= 0:
                            cl = chr(ord('A') + (n % 26)) + cl
                            n = n // 26 - 1
                        marker = ">>>" if c == col_idx else "   "
                        print(f"{marker} {cl}: {row[c][:MAX_PREVIEW] if c < len(row) else ''}")

                    if total_matches >= max_matches:
                        print(f"\n--- Reached match limit ({max_matches}), stopping ---")
                        return total_matches

    return total_matches


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Search within spreadsheet cells")
    parser.add_argument("path", help="Path to spreadsheet file")
    parser.add_argument("query", help="Search query (regex or literal string)")
    parser.add_argument("--context", type=int, default=2,
                        help="Context columns around match (default: 2)")
    parser.add_argument("--literal", action="store_true",
                        help="Treat query as literal string")
    parser.add_argument("--max-matches", type=int, default=50,
                        help="Maximum matches to report (default: 50)")
    args = parser.parse_args()

    # Compile pattern
    if args.literal:
        pattern = re.compile(re.escape(args.query), re.IGNORECASE)
    else:
        try:
            pattern = re.compile(args.query, re.IGNORECASE)
        except re.error as e:
            print(f"Invalid regex: {e}", file=sys.stderr)
            sys.exit(1)

    ext = os.path.splitext(args.path)[1].lower()

    if ext in ('.csv', '.tsv'):
        total = search_csv(args.path, pattern, args.context, args.max_matches)
    elif ext in ('.xlsx', '.xlsm'):
        total = search_xlsx(args.path, pattern, args.context, args.max_matches)
    elif ext == '.xls':
        total = search_xls(args.path, pattern, args.context, args.max_matches)
    else:
        print(f"ERROR: Unsupported format: {ext}", file=sys.stderr)
        sys.exit(1)

    if total == 0:
        print(f"No matches found for: {args.query}")
    else:
        print(f"\n--- {total} match(es) found ---")
